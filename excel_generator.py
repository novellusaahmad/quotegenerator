"""
Excel Generator for Novellus Loan Management System
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import tempfile
import os
from datetime import datetime
from dateutil.relativedelta import relativedelta

class NovellussExcelGenerator:
    """Excel generator for loan quotes and calculations"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
    
    def generate_quote_excel(self, quote_data, application_data=None):
        """Generate Excel quote document"""
        # Create new workbook
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Loan Quote"
        
        # Set up styles
        title_font = Font(name='Arial', size=16, bold=True)
        header_font = Font(name='Arial', size=12, bold=True)
        normal_font = Font(name='Arial', size=10)
        
        # Add title
        self.worksheet['A1'] = "Novellus Finance - Loan Quote"
        self.worksheet['A1'].font = title_font
        self.worksheet['A1'].alignment = Alignment(horizontal='center')
        self.worksheet.merge_cells('A1:B1')
        
        # Add quote details
        row = 3
        details = [
            ('Quote ID:', str(quote_data.get('id', 'N/A'))),
            ('Date:', datetime.now().strftime('%d/%m/%Y')),
            ('Gross Amount:', f"£{quote_data.get('gross_amount', 0):,.2f}"),
            ('Net Advance:', f"£{quote_data.get('net_advance', 0):,.2f}"),
            ('Interest Rate:', f"{quote_data.get('interest_rate', 0):.2f}%"),
            ('Loan Term:', f"{quote_data.get('loan_term', 0)} months"),
            ('Monthly Payment:', f"£{quote_data.get('monthly_payment', 0):,.2f}"),
            ('Total Interest:', f"£{quote_data.get('total_interest', 0):,.2f}"),
        ]
        
        for key, value in details:
            self.worksheet[f'A{row}'] = key
            self.worksheet[f'A{row}'].font = header_font
            self.worksheet[f'B{row}'] = value
            self.worksheet[f'B{row}'].font = normal_font
            row += 1
        
        # Auto-adjust column widths
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            self.workbook.save(tmp_file.name)
            tmp_path = tmp_file.name
        
        # Read the content
        with open(tmp_path, 'rb') as f:
            excel_content = f.read()
        
        # Clean up
        os.unlink(tmp_path)

        return excel_content

    def generate_detailed_schedules_excel(self, payment_schedule, tranche_schedule, params):
        """Generate Excel workbook with payment and tranche schedules.

        The workbook includes raw value sheets and companion sheets that use
        formulas so results can be recalculated if inputs like dates or
        tranche amounts change in Excel."""

        # --- Parameter sheet ---
        self.workbook = openpyxl.Workbook()
        params_ws = self.workbook.active
        params_ws.title = "Parameters"

        annual_rate = float(params.get("annual_rate", 0)) / 100.0
        start_date_str = params.get("start_date")
        loan_term = int(params.get("loan_term", len(payment_schedule)))
        use_360 = params.get("use_360_days", False)
        days_in_year = 360 if use_360 else 365

        if start_date_str:
            start_dt = datetime.strptime(start_date_str, "%Y-%m-%d")
        else:
            start_dt = datetime.today()
        loan_end_dt = start_dt + relativedelta(months=loan_term)

        initial_balance = 0.0
        if payment_schedule:
            initial_balance = float(payment_schedule[0].get("opening_balance", 0))

        params_ws["A1"], params_ws["B1"] = "Annual Rate", annual_rate
        params_ws["A2"], params_ws["B2"] = "Periodic Rate", "=B1/12"
        params_ws["A3"], params_ws["B3"] = "Days In Year", days_in_year
        params_ws["A4"], params_ws["B4"] = "Initial Balance", initial_balance
        params_ws["A5"], params_ws["B5"] = "Start Date", start_dt
        params_ws["A6"], params_ws["B6"] = "Loan End Date", loan_end_dt

        params_ws["B1"].number_format = "0.00%"
        params_ws["B2"].number_format = "0.00%"
        params_ws["B5"].number_format = "dd/mm/yyyy"
        params_ws["B6"].number_format = "dd/mm/yyyy"

        # --- Payment schedule (values) ---
        pay_ws = self.workbook.create_sheet("Payment Schedule")
        pay_headers = ["Payment #", "Date", "Opening Balance", "Payment", "Interest", "Closing Balance"]
        for col, header in enumerate(pay_headers, 1):
            pay_ws.cell(row=1, column=col, value=header)

        for r, payment in enumerate(payment_schedule, start=2):
            pay_ws.cell(row=r, column=1, value=r - 1)
            pay_ws.cell(row=r, column=2, value=payment.get("date"))
            pay_ws.cell(row=r, column=3, value=float(payment.get("opening_balance", 0)))
            pay_ws.cell(row=r, column=4, value=float(payment.get("payment_amount", 0)))
            pay_ws.cell(row=r, column=5, value=float(payment.get("interest_amount", 0)))
            pay_ws.cell(row=r, column=6, value=float(payment.get("closing_balance", 0)))

        for col in (3, 4, 5, 6):
            for r in range(2, len(payment_schedule) + 2):
                pay_ws.cell(row=r, column=col).number_format = "£#,##0.00"

        # --- Payment schedule (formulas) ---
        pay_form_ws = self.workbook.create_sheet("Payment Schedule Data")
        for col, header in enumerate(pay_headers, 1):
            pay_form_ws.cell(row=1, column=col, value=header)

        for r in range(2, len(payment_schedule) + 2):
            pay_form_ws.cell(row=r, column=1, value="=ROW()-1")
            pay_form_ws.cell(row=r, column=2, value=f"='Payment Schedule'!B{r}")
            if r == 2:
                pay_form_ws.cell(row=r, column=3, value="=Parameters!$B$4")
            else:
                pay_form_ws.cell(row=r, column=3, value=f"=F{r-1}")
            pay_form_ws.cell(row=r, column=4, value=f"='Payment Schedule'!D{r}")
            pay_form_ws.cell(row=r, column=5, value=f"=C{r}*Parameters!$B$2")
            pay_form_ws.cell(row=r, column=6, value=f"=C{r}-D{r}+E{r}")

        for col in (3, 4, 5, 6):
            for r in range(2, len(payment_schedule) + 2):
                pay_form_ws.cell(row=r, column=col).number_format = "£#,##0.00"

        # --- Tranche schedule (values) ---
        tranche_ws = self.workbook.create_sheet("Tranche Schedule")
        tranche_headers = ["Tranche #", "Release Date", "Amount", "Days Outstanding", "Rate", "Interest"]
        for col, header in enumerate(tranche_headers, 1):
            tranche_ws.cell(row=1, column=col, value=header)

        for r, tranche in enumerate(tranche_schedule, start=2):
            release = tranche.get("release_date") or tranche.get("date")
            days_outstanding = 0
            if release:
                try:
                    rel_dt = datetime.strptime(release, "%Y-%m-%d")
                    days_outstanding = (loan_end_dt - rel_dt).days
                except Exception:
                    release = tranche.get("release_date", "")

            amount = float(tranche.get("amount", 0))
            rate = float(tranche.get("interest_rate", params.get("annual_rate", 0))) / 100.0
            interest = amount * rate * days_outstanding / days_in_year if days_outstanding else 0

            tranche_ws.cell(row=r, column=1, value=tranche.get("tranche_number", r - 1))
            tranche_ws.cell(row=r, column=2, value=release)
            tranche_ws.cell(row=r, column=3, value=amount)
            tranche_ws.cell(row=r, column=4, value=days_outstanding)
            tranche_ws.cell(row=r, column=5, value=rate)
            tranche_ws.cell(row=r, column=6, value=interest)

        for r in range(2, len(tranche_schedule) + 2):
            tranche_ws.cell(row=r, column=3).number_format = "£#,##0.00"
            tranche_ws.cell(row=r, column=5).number_format = "0.00%"
            tranche_ws.cell(row=r, column=6).number_format = "£#,##0.00"

        # --- Tranche schedule (formulas) ---
        tranche_form_ws = self.workbook.create_sheet("Tranche Schedule Data")
        for col, header in enumerate(tranche_headers, 1):
            tranche_form_ws.cell(row=1, column=col, value=header)

        for r in range(2, len(tranche_schedule) + 2):
            tranche_form_ws.cell(row=r, column=1, value="=ROW()-1")
            tranche_form_ws.cell(row=r, column=2, value=f"='Tranche Schedule'!B{r}")
            tranche_form_ws.cell(row=r, column=3, value=f"='Tranche Schedule'!C{r}")
            tranche_form_ws.cell(row=r, column=4, value="=Parameters!$B$6-B{r}")
            tranche_form_ws.cell(row=r, column=5, value=f"='Tranche Schedule'!E{r}")
            tranche_form_ws.cell(row=r, column=6, value=f"=C{r}*E{r}*D{r}/Parameters!$B$3")

        for r in range(2, len(tranche_schedule) + 2):
            tranche_form_ws.cell(row=r, column=3).number_format = "£#,##0.00"
            tranche_form_ws.cell(row=r, column=5).number_format = "0.00%"
            tranche_form_ws.cell(row=r, column=6).number_format = "£#,##0.00"

        # Save workbook to temporary file and return bytes
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            self.workbook.save(tmp_file.name)
            tmp_path = tmp_file.name

        with open(tmp_path, "rb") as f:
            excel_bytes = f.read()

        os.unlink(tmp_path)

        return excel_bytes
    
    def generate_payment_schedule_excel(self, payment_schedule, quote_data):
        """Generate Excel payment schedule"""
        # Create new workbook
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Payment Schedule"
        
        # Set up styles
        title_font = Font(name='Arial', size=16, bold=True)
        header_font = Font(name='Arial', size=12, bold=True)
        normal_font = Font(name='Arial', size=10)
        
        # Add title
        self.worksheet['A1'] = "Novellus Finance - Payment Schedule"
        self.worksheet['A1'].font = title_font
        self.worksheet['A1'].alignment = Alignment(horizontal='center')
        self.worksheet.merge_cells('A1:F1')
        
        # Add headers
        headers = ['Payment #', 'Date', 'Opening Balance', 'Payment', 'Interest', 'Closing Balance']
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Add payment schedule data
        row = 4
        for i, payment in enumerate(payment_schedule, 1):
            self.worksheet.cell(row=row, column=1).value = i
            self.worksheet.cell(row=row, column=2).value = payment.get('date', '')
            self.worksheet.cell(row=row, column=3).value = f"£{payment.get('opening_balance', 0):,.2f}"
            self.worksheet.cell(row=row, column=4).value = f"£{payment.get('payment_amount', 0):,.2f}"
            self.worksheet.cell(row=row, column=5).value = f"£{payment.get('interest_amount', 0):,.2f}"
            self.worksheet.cell(row=row, column=6).value = f"£{payment.get('closing_balance', 0):,.2f}"
            row += 1
        
        # Auto-adjust column widths
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            self.workbook.save(tmp_file.name)
            tmp_path = tmp_file.name
        
        # Read the content
        with open(tmp_path, 'rb') as f:
            excel_content = f.read()
        
        # Clean up
        os.unlink(tmp_path)
        
        return excel_content